import os
import shutil
import json
import csv
from datetime import datetime

from docx import Document              # python-docx
from openpyxl import load_workbook
from pdfminer.high_level import extract_text as pdf_extract_text

# ------------------------------
# Config
# ------------------------------
SCAN_FOLDER = input("Enter folder to scan: ").strip().strip('"')
OUTPUT_FOLDER = "FileGrabber"  # destination root
KEYWORDS = ["password", "secret", "token", "key", "api"]
MAX_FILE_SIZE_MB = None  # None = no size limit, or set int like 500

TEXT_EXTS = {
    ".txt", ".log", ".csv", ".json", ".md", ".ini", ".cfg", ".conf",
    ".xml", ".yml", ".yaml", ".env"
}

def bytes_keyword_hits(path, keywords, chunk_size=1024 * 1024):
    """Scan arbitrary binary file in chunks; return hit offsets (approx)."""
    hits = []
    try:
        # pre encode keywords as ascii/utf-8 lowercase byte sets
        kw_bytes = [k.encode("utf-8").lower() for k in keywords]
        with open(path, "rb") as f:
            offset = 0
            while True:
                chunk = f.read(chunk_size)
                if not chunk:
                    break
                cl = chunk.lower()
                for kb in kw_bytes:
                    idx = cl.find(kb)
                    if idx != -1:
                        hits.append({"offset": int(offset + idx), "keyword": kb.decode("utf-8")})
                offset += len(chunk)
    except Exception:
        pass
    return hits

def text_keyword_matches(text, keywords):
    matches = []
    try:
        lines = text.splitlines()
        for i, line in enumerate(lines, 1):
            low = line.lower()
            for kw in keywords:
                if kw in low:
                    matches.append({"line_number": i, "keyword": kw, "line": line.strip()})
    except Exception:
        pass
    return matches

def read_text_file(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception:
        try:
            with open(path, "r", encoding="latin-1", errors="ignore") as f:
                return f.read()
        except Exception:
            return ""

def read_docx_text(path):
    try:
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception:
        return ""

def read_xlsx_text(path, max_rows_per_sheet=100000):
    out = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            rows = 0
            for row in sheet.iter_rows(values_only=True):
                rows += 1
                # stop runaway sheets
                if rows > max_rows_per_sheet:
                    break
                vals = [str(c) for c in row if c is not None]
                if vals:
                    out.append(" ".join(vals))
    except Exception:
        return ""
    return "\n".join(out)

def read_pdf_text(path):
    try:
        return pdf_extract_text(path) or ""
    except Exception:
        return ""

def extract_text_by_ext(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".docx":
        return read_docx_text(path)
    if ext == ".xlsx":
        return read_xlsx_text(path)
    if ext == ".pdf":
        return read_pdf_text(path)
    if ext in TEXT_EXTS:
        return read_text_file(path)
    return None

def size_mb(path):
    try:
        return os.path.getsize(path) / (1024 * 1024)
    except Exception:
        return 0.0

def copy_preserving_tree(src_path, base_src, base_dst):
    rel_path = os.path.relpath(src_path, base_src)
    dst_path = os.path.join(base_dst, rel_path)
    os.makedirs(os.path.dirname(dst_path), exist_ok=True)
    try:
        shutil.copy2(src_path, dst_path)
        return dst_path, True, ""
    except Exception as e:
        return dst_path, False, str(e)

def main():
    if not os.path.isdir(SCAN_FOLDER):
        print("Provided path is not a directory.")
        return

    if os.path.exists(OUTPUT_FOLDER):
        shutil.rmtree(OUTPUT_FOLDER, ignore_errors=True)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    log = []
    total_files = 0

    for root, dirs, files in os.walk(SCAN_FOLDER):
        for name in files:
            total_files += 1
            fpath = os.path.join(root, name)

            if MAX_FILE_SIZE_MB is not None:
                try:
                    if size_mb(fpath) > MAX_FILE_SIZE_MB:
                        log.append({
                            "original_path": fpath,
                            "copied": False,
                            "skip_reason": f"size>{MAX_FILE_SIZE_MB}MB",
                            "size_bytes": os.path.getsize(fpath),
                            "matches": [],
                            "binary_hits": [],
                            "extractor": None,
                        })
                        continue
                except Exception:
                    pass

            ext_text = extract_text_by_ext(fpath)
            matches = []
            binary_hits = []

            if ext_text is None:
                # Unknown/binary extension → do raw byte scan for keywords
                binary_hits = bytes_keyword_hits(fpath, [k.lower() for k in KEYWORDS])
            else:
                # We have text → search with line numbers
                matches = text_keyword_matches(ext_text, [k.lower() for k in KEYWORDS])

            dst, copied, err = copy_preserving_tree(fpath, SCAN_FOLDER, OUTPUT_FOLDER)

            entry = {
                "original_path": fpath,
                "copied_path": dst,
                "copied": copied,
                "copy_error": err,
                "size_bytes": os.path.getsize(fpath) if os.path.exists(fpath) else None,
                "matches": matches,
                "binary_hits": binary_hits,
                "extractor": (
                    "docx" if fpath.lower().endswith(".docx") else
                    "xlsx" if fpath.lower().endswith(".xlsx") else
                    "pdf"  if fpath.lower().endswith(".pdf")  else
                    "text" if ext_text is not None and fpath.lower().split(".")[-1] in {e.strip(".") for e in TEXT_EXTS} else
                    None
                ),
            }
            log.append(entry)

    meta = {
        "scanned_root": os.path.abspath(SCAN_FOLDER),
        "output_root": os.path.abspath(OUTPUT_FOLDER),
        "total_files_seen": total_files,
        "keywords": KEYWORDS,
        "timestamp": datetime.utcnow().isoformat() + "Z",
    }
    with open(os.path.join(OUTPUT_FOLDER, "log.json"), "w", encoding="utf-8") as f:
        json.dump({"meta": meta, "items": log}, f, ensure_ascii=False, indent=2)

    with open(os.path.join(OUTPUT_FOLDER, "log.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Original Path", "Copied", "Copy Error", "Size Bytes", "Extractor", "Text Matches", "Binary Hits"])
        for it in log:
            text_hits_str = "; ".join(f"{m['keyword']}@{m.get('line_number','?')}" for m in it["matches"]) if it["matches"] else ""
            bin_hits_str = "; ".join(f"{h['keyword']}@{h['offset']}" for h in it["binary_hits"]) if it["binary_hits"] else ""
            w.writerow([
                it["original_path"],
                it["copied"],
                it["copy_error"],
                it["size_bytes"],
                it["extractor"],
                text_hits_str,
                bin_hits_str
            ])

    print(f"Scanning complete. Copied tree: '{OUTPUT_FOLDER}'. "
          f"Logs: log.json, log.csv")

if __name__ == "__main__":
    main()
